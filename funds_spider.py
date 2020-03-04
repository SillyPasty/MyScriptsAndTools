import re
import traceback
import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from openpyxl import load_workbook


def get_html(url):
    try:
        r = requests.get(url)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return r.text
    except:
        return ""


def get_url_list():
    '''return url list of all funds'''
    fund_list_url = 'http://fund.eastmoney.com/allfund.html'
    text = get_html(fund_list_url)
    fund_url_list = re.findall(r'http://fund.eastmoney.com/\d{6}.html', text)
    return fund_url_list


def get_fund_info(fund_url):
    '''receive fund url, return a fund info dict'''
    fund_info = {}
    fund_info['id'] = re.findall(r'\d{6}', fund_url)[0]

    text = get_html(fund_url)
    if text is "":
        print(fund_url)
        return ""
    soup = BeautifulSoup(text, 'html.parser')

    fund_info_table = soup.select('.infoOfFund')[0].find('table').find_all('td')

    info_t_a = fund_info_table[0].find('a')
    if info_t_a is None:
        return None
    fund_info['type'] = fund_info_table[0].find('a').text
    if fund_info['type'] == '货币型' or fund_info['type'] == '混合-FOF':
        return None

    fund_info['name'] = soup.find('span', attrs={'class':'funCur-FundName'}).text

    risk = re.findall(r'\S{1,2}风险', fund_info_table[0].text)
    if len(risk):
        fund_info['risk'] = risk[0]
    else:
        fund_info['risk'] = ''
    fund_info['scale'] = fund_info_table[1].text.split('：')[1].split('（')[0]

    fund_return_list = soup.find_all('div', attrs={'class':'Rdata'})[:8]
    fund_return_dict = {}
    returns_list = ['_', '1month', '3month', '6month', '_', '1year', '2year', '3year']
    for idx, item in enumerate(fund_return_list):
        if idx != 0 and idx != 4:
            fund_return_dict[returns_list[idx]] = item.text
    fund_info['returns'] = fund_return_dict

    fund_rank_dict = {}
    fund_rank_info_list = soup.find_all('div', attrs={'class':'Rdata'})[24:32]
    if not re.match(fund_rank_info_list[1].text, '|'):
        fund_rank_info_list = soup.find_all('div', attrs={'class':'Rdata'})[32:40]
    for idx, item in enumerate(fund_rank_info_list):
        if idx != 0 and idx != 4:
            fund_rank_dict[returns_list[idx]] = item.text

    fund_info['rank'] = fund_rank_dict

    fund_manager_name_l = fund_info_table[2].find_all('a')
    fund_manager_name_list = []
    fund_manager_url_list = []
    for fund_manager_name in fund_manager_name_l:
        fund_manager_name_list.append(fund_manager_name.text)
        fund_manager_url_list.append(fund_manager_name.attrs['href'])

    fund_info['manager_list'] = fund_manager_name_list
    fund_info['establish_date'] = fund_info_table[3].text.split('：')[1]
    if fund_info['establish_date'] != '--':
        fund_info['establish_length'] = int((datetime.now().date() - datetime.strptime(fund_info['establish_date'], '%Y-%m-%d').date()).days / 365)
    else:
        fund_info['establish_length'] = '--'
    fund_info['company'] = fund_info_table[4].text.split('：')[1]
    if len(fund_info_table) > 6:
        tmp = fund_info_table[6].text.split('：')
        fund_info['follow'] = tmp[1].split('|')[0]
        fund_info['follow_error'] = tmp[2]
    else:
        fund_info['follow'] = ''
        fund_info['follow_error'] = ''


    fund_manager_url = fund_manager_url_list[0]
    text_manager = get_html(fund_manager_url)
    soup_manager = BeautifulSoup(text_manager, 'html.parser')
    manager_info_table = soup_manager.select('.space8')[1].next_sibling.find_all('td')

    fund_info['manager_start'] = manager_info_table[0].text
    # fund_info['manager_end'] = manager_info_table[1].text
    fund_info['manager_period'] = manager_info_table[3].text
    fund_info['manager_returns'] = manager_info_table[4].text

    

    fund_important_stock_t = soup.find_all('table', attrs={'class': 'ui-table-hover'})[0]
    fund_imp = fund_important_stock_t.find_all('a')
    fund_important_stock_list = []
    for i in fund_imp:
        if i.text != '股吧':
            fund_important_stock_list.append(i.text)
            fund_important_stock_list.append(i.parent.next_sibling.next_sibling.text)
    fund_info['big stock position'] = fund_important_stock_list

    fund_important_bond_t = soup.find_all('table', attrs={'class': 'ui-table-hover'})[1]
    fund_imp = fund_important_bond_t.find_all('a')
    fund_important_bond_list = []
    for i in fund_imp:
        fund_important_bond_list.append(i.text)
        fund_important_bond_list.append(i.parent.next_sibling.next_sibling.text)
    fund_info['big bond position'] = fund_important_bond_list
    
    return fund_info


def get_fund_star(fund_id):
    fund_star = {'3 years': '', '5 years': ''}
    return fund_star 


def get_manager_info(manager_name):
    manager_info = {}
    return manager_info


def write_fund_info(fund_info, wb):
    title_list = ['id', 'name', 'type', 'risk', 'establish_date', 'establish_length', 'scale', 'company', 'follow', 'follow_error', 'manager_list', 'manager_start', 'manager_period', 'manager_returns', 
                  'returns', 'big stock position', 'big bond position']
    returns_list = ['3month', '6month', '1year', '2year', '3year']
    fund_list_sheet = wb.get_sheet_by_name('Funds')
    max_row = fund_list_sheet.max_row
    info_list = []
    for key in title_list:
        if key == 'returns':
            for return_key in returns_list:
                info_list.append(fund_info[key][return_key])
                info_list.append(fund_info['rank'][return_key])
        elif key == 'big stock position' or key == 'big bond position':
            i = 0
            for item in fund_info[key]:
                i += 1
                info_list.append(item)
            while i < 10:
                i += 1
                info_list.append('')
        elif key == 'manager_list':
            info_list.append(fund_info[key][0])
        else:
            info_list.append(fund_info[key])
    fund_list_sheet.append(info_list)
    # wb.save(path)
    # print('success')
    return


def main():
    excel_path = r'D:/test.xlsx'
    url_list_path = r'D:/url.txt'
    wb = load_workbook(excel_path)
    url_list = []
    PER = 0.5127
    print('Getting url list...')
    # url_list = get_url_list()
    for line in open(url_list_path, 'r'):
        url_list.append(line[:-1])
    print('Done.')

    # with open(url_list_path, 'w') as f:
    #     for url in url_list:
    #          f.write(url)
    #          f.write('\n')

    length = len(url_list)
    print(str(length) + ' items total.')
    print('Searching and writing...')
    # tem = length
    # tem = int(length - length * PER)

    for idx, url in enumerate(reversed(url_list)):
        if idx % 1000 == 1:
            wb = load_workbook(excel_path)
        info = get_fund_info(url)
        if info is None:
            continue
        write_fund_info(info, wb)
        print('\r当前进度: {:.2f}% 链接:{}'.format(idx*100/length, url), end='')
        if idx % 1000 == 0:
            wb.save(excel_path)
    wb.save(excel_path)
    print('success')

# print(get_fund_info('http://fund.eastmoney.com/000008.html'))

def test():
    print('----\n')
    print(get_fund_info('http://fund.eastmoney.com/005911.html'))
# main()
test()